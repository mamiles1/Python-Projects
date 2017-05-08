from bs4 import BeautifulSoup
from bs4 import SoupStrainer
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
import openpyxl
import pyautogui, time
import re

wb = openpyxl.load_workbook('fullbook skus with UPC.xlsx')
wb.get_sheet_names()
MasterSheet= wb.get_sheet_by_name('Sheet1')

upcHolder=[]

# appends data from excel sheet into an array holder
for rowcellObj in MasterSheet['D2':MasterSheet.max_row]:

    for cellObj in rowcellObj:
        upcHolder.append(str(cellObj.value).zfill(12))

print("---Loaded UPCs from excel sheet into array---")

skuNameHolder=[]
skuPriceHolder=[]

browser = webdriver.PhantomJS()
browser.get('https://www.walmart.com/')
browser.set_window_size(1120,550)
wait = WebDriverWait(browser,5)

upcPosition=0

while upcPosition <=len(upcHolder):
    # find the search box to enter upc on website
    searchfinder=browser.find_element_by_xpath('//*[@id="global-search-input"]')
    searchfinder.click()
    time.sleep(1)
    # enters upc into search box and clicks to find upc
    searchfinder.send_keys(upcHolder[upcPosition])
    searchfinder.submit()

    print(upcPosition)
    print(upcHolder[upcPosition])
    wait

    htmlcode = browser.page_source
    soupPull = BeautifulSoup(htmlcode, "html.parser")
    soup_string=str(soupPull)

    matches = re.findall("product-title-link line-clamp line-clamp-3", soup_string)

    # searches if array upc is available on website
    if(matches):
        # if upc is found on website name and price is entered into array holder
        for ItemName in soupPull.find_all('a',{"product-title-link line-clamp line-clamp-3"},limit=1):

            skuNameHolder.append(ItemName.text)
            print(ItemName.text)

        for offerPrice in soupPull.find_all('span',{"display-inline-block arrange-fit Price u-textColor price-main"}):

            skuPriceHolder.append(offerPrice.text)
            print(offerPrice.text)

    else:
        skuNameHolder.append("Not Found")
        skuPriceHolder.append("Not Found")

    time.sleep(4)
    #deletes array from search box
    searchfinder = browser.find_element_by_xpath('//*[@id="global-search-input"]')
    searchfinder.click()
    time.sleep(6)
    #pyautogui.hotkey('ctrl', 'a')
    wait
    searchfinder.clear()
    #pyautogui.typewrite(['backspace', 'backspace', 'backspace', 'backspace', 'backspace', 'backspace', 'backspace','backspace','backspace','backspace','backspace','backspace','backspace'])
    upcPosition+=1
    time.sleep(4)

namecount=2
pricecount=2
# places upcs in array into excel sheet
for num in skuPriceHolder:
    MasterSheet.cell(row=namecount, column=5).value = num
    pricecount+=1

for num in skuPriceHolder:
    MasterSheet.cell(row=pricecount, column=6).value = num
    pricecount+=1

wb.save('fullbook skus with UPC.xlsx')


