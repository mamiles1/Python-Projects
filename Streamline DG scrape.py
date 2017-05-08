from bs4 import BeautifulSoup
from bs4 import SoupStrainer
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
import openpyxl


wb = openpyxl.load_workbook('DGtest1.xlsx')
Mainsheet=wb.get_sheet_by_name('Sheet1')

# request to dollar general main page
browser = webdriver.PhantomJS()
url = browser.get('http://www.dollargeneral.com')
browser.set_window_size(1120,550)

wait = WebDriverWait(browser,2)

buttoncounter=3
secondfinder=1
nextpagefinder=1
secondbutton=1

SKUname = []
SKUprice = []
oldprice = []
SKUfinalID = []
catlist=[]

while True:
    #find the category from the menu page
    try:
        oldurl = browser.current_url
        shopbutton=browser.find_element_by_xpath('//*[@id="nav_menu"]/li/a')
        shopbutton.click()
        catbutton = browser.find_elements_by_xpath('//*[@id="department"]/li['+str(buttoncounter)+']/a')

        for pushbutton in catbutton:
            print(pushbutton.get_attribute("href"))

        pushbutton.click()
        wait
        buttoncounter+=1
        secondfinder=1


        while True:
            try:
                try:
                    # clicks on view all buttons after finding html links
                    nextlevelclick = browser.find_element_by_xpath('//*[@id="maincontent"]/div[5]/div[1]/div[7]/dd/ol/ul[' + str(secondfinder) + ']/div/a')

                    print("secondfinder ")
                    print(secondfinder)
                    nextlevelclick.click()
                    secondfinder += 1
                    wait

                except:
                    nextlevelclick=browser.find_element_by_xpath('// *[ @ id = "maincontent"] / div[5] / div[1] / div[7] / dd / ol / ul['+str(secondfinder)+'] / li[1] / a')
                    print("except secondfinder ")
                    print(secondfinder)
                    nextlevelclick.click()
                    secondfinder += 1
                    wait
                secondbutton=1
                while True:
                    try:
                        deepclick = browser.find_element_by_xpath('//*[@id="maincontent"]/div[5]/div[1]/div[7]/dd/ol/ul['+str(secondbutton)+']/div/a')

                        deepclick.click()
                        print("secondbutton ")
                        print(secondbutton)
                        secondbutton+=1

                    except:
                        pass

                    # Grabs data from webpage code
                    try:

                        NextHtml = browser.page_source
                        noBestSeller = SoupStrainer(class_="columns")
                        internalsoup = BeautifulSoup(NextHtml, "html.parser", parse_only=noBestSeller)

                        for freshlink in internalsoup.find_all('a', {"class": "product-item-link"}):
                            SKUname.append(freshlink.text)
                            catnamefinder = browser.find_element_by_xpath('// *[ @ id = "page-title-heading"] / span')
                            catlist.append(catnamefinder.text)

                        for freshlink2 in internalsoup.find_all('span',
                                                                class_="price-container price-final_price tax weee"):

                            if "As low as" in freshlink2.text or "Was" in freshlink2.text:
                                continue
                            else:
                                SKUprice.append(freshlink2.text)
                                SKUfinalID.append(freshlink2.span.get('id'))

                        for freshlink3 in internalsoup.find_all('span',
                                                                class_="price-container price-final_price tax weee"):
                            if "As low as" in freshlink3.text or "Was" in freshlink3.text:
                                oldprice.append(freshlink3.text)

                        SKUname = [clean.strip() for clean in SKUname]

                    except:
                        continue

                    try:
                        nextpage = browser.find_element_by_css_selector(
                            '#maincontent > div.columns > div.column.main > div.product-list-bar > div:nth-child(3) > div.pages > ul > li.item.pages-item-next > a')
                        nextpage.click()

                    except:
                        try:
                            secondarybackpage = browser.find_element_by_xpath('/html/body/div[1]/div[3]/ul/li[3]/a')
                            secondarybackpage.click()
                            wait
                            continue

                        except:
                            backpage = browser.find_element_by_xpath('/html/body/div[1]/div[3]/ul/li[2]/a')
                            backpage.click()
                            wait
                            break

            except:
                break

    except:
        break

t = 2
r = 2
z = 2
y = 2
catcount=2

# places website data into excel file
for num in catlist:
    Mainsheet.cell(row=catcount, column=1).value = num
    catcount+=1

for num in SKUname:
    Mainsheet.cell(row=r, column=2).value = num
    r += 1

for num in SKUprice:
    Mainsheet.cell(row=t, column=3).value = num
    t += 1

for num in SKUfinalID:
    Mainsheet.cell(row=z, column=4).value = num
    z += 1

for num in oldprice:
    Mainsheet.cell(row=y, column=5).value = num
    y += 1

wb.save('DGtest1.xlsx')






